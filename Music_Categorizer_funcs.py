import re
import os
import openpyxl
import eyed3
folder_path = "C:\\Users\\MT1ShotYT\\Desktop\\music"


def all_files_list(folder_path):
    mpath = folder_path
    files = [f for f in os.listdir(mpath)]
    return files


def music_artists_name_list_finder(folder_path):
    folder = folder_path
    files = os.listdir(folder)
    artists_names_list = []
    for x in files:
        if 'mp3' in x:
            mp3 = eyed3.load(folder + '/' + x)
            nam = mp3.tag.artist  
            artists_names_list.append(nam)
        else:
            pass
    return artists_names_list


def music_name_finder(files):
    artists_names_list = []
    for musicname in files:
        if len(musicname) > 0 :
            x = musicname.split('-')
            c = ''
            counter = 0
            for i in x[0]:
                if i == '.':
                    break
                else:
                    if i.isalpha() or i == ' ':
                        if i == ' ' and counter <= 1:
                            pass
                        else:    
                            c += i
                            counter += 1
                            if counter >= 30:
                                break
                    else:
                        continue
            artists_names_list.append(c)
    return artists_names_list
                

def folder_creation(path, folder_name):
    main_dir = path + '\\' + folder_name
    if not os.path.exists(main_dir):
        os.mkdir(main_dir)
        print(f"Created folder '{main_dir}'.")
    else:
        # print(f"Folder '{main_dir}' already exists.")
        pass
        
        
def all_music_fullname_finder(folder_path):
    files = all_files_list(folder_path)
    lst = music_name_finder(files)
    for i in range(len(lst)):
        for j in  range(len(files)):
            if lst[i] in files[j]:
                print(lst[i])
                print(files[j])
                
                
def music_fullname_finder(folder_path, artist_name):         
    folder = folder_path
    files = os.listdir(folder)
    for x in files:
        if 'mp3' in x:
            mp3 = eyed3.load(folder + '/' + x)
            if mp3.tag.artist == artist_name:
                return x
            else:
                pass
        else:
            pass
            

def path_definer(movefrom, moveto, music_name):
    movefrom = movefrom + '\\' + music_name
    moveto = moveto + '\\' + music_name
    
    return movefrom, moveto
    
            
def move_music_to_directory(folder_path, initfoldername, music_name):
    movefrom = folder_path + '\\' + music_name
    moveto = folder_path + '\\' + initfoldername + '\\' +music_name
    location = folder_path + '\\' + initfoldername 
    os.rename(movefrom, moveto)
    print(f'file: {music_name} has been moved to folder: {location}')
    

def DB_Music_Name_Check(db_file_name, artist_name):
    theFile = openpyxl.load_workbook(db_file_name)
    allSheetNames = theFile.sheetnames

    for x in allSheetNames:
        currentSheet = theFile[x]
        count = 0
        for i in range(1, currentSheet.max_row+1):
            sheet = f'A{i}'
            count += 1
            if artist_name in currentSheet[sheet].value:
                return True
            elif count == currentSheet.max_row+1:
                return False
            else:
                continue
            
            
def music_categorizer(folder_path, db_file_name, folder_name):
    artistname = music_artists_name_list_finder(folder_path)
    for artname in artistname:
        existence = DB_Music_Name_Check(db_file_name, artist_name=artname)  
        if existence is True:
            musicfullname = music_fullname_finder(folder_path, artist_name=artname)
            folder_creation(path=folder_path, folder_name=folder_name)
            move_music_to_directory(folder_path, initfoldername= folder_name, music_name=musicfullname)        
        elif existence == None:
            pass
            # print("No matching info found :(") 
        else:
            pass
    
